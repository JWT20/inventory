import csv
import io
import logging
import os
import uuid

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from sqlalchemy.orm import Session

from app.auth import get_current_user, require_product_manager
from app.config import settings
from app.database import get_db
from app.events import publish_event
from app.models import SKU, Order, OrderLine, User
from app.routers.skus import _sku_to_response, generate_display_name, generate_sku_code
from app.schemas import (
    OrderImportResult,
    OrderLineResponse,
    OrderResponse,
    ScanResult,
    SKUResponse,
)
from app.services.embedding import process_image
from app.services.matching import find_best_matches

logger = logging.getLogger(__name__)
router = APIRouter(
    prefix="/orders", tags=["orders"], dependencies=[Depends(get_current_user)]
)


def _line_to_response(line: OrderLine) -> OrderLineResponse:
    return OrderLineResponse(
        id=line.id,
        sku_id=line.sku_id,
        sku_code=line.sku.sku_code,
        sku_name=line.sku.name,
        quantity=line.quantity,
        scanned_quantity=line.scanned_quantity,
    )


def _order_to_response(order: Order) -> OrderResponse:
    return OrderResponse(
        id=order.id,
        order_number=order.order_number,
        customer_name=order.customer_name,
        status=order.status,
        created_at=order.created_at,
        updated_at=order.updated_at,
        lines=[_line_to_response(l) for l in order.lines],
    )


def _parse_csv(content: bytes) -> list[dict]:
    """Parse CSV content. Expected columns: producent, wijnnaam, type, jaargang, volume, aantal."""
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text), delimiter=",")

    # Normalize header names (strip whitespace, lowercase)
    if reader.fieldnames:
        reader.fieldnames = [f.strip().lower() for f in reader.fieldnames]

    rows = []
    for i, row in enumerate(reader, start=2):
        row = {k.strip().lower(): v.strip() for k, v in row.items() if k}
        producent = row.get("producent", "").strip()
        wijnnaam = row.get("wijnnaam", "").strip()
        wtype = row.get("type", "").strip()
        jaargang = row.get("jaargang", "").strip()
        volume = row.get("volume", "0.75L").strip()
        aantal = row.get("aantal", "1").strip()

        if not producent or not wijnnaam:
            raise HTTPException(400, f"Rij {i}: 'producent' en 'wijnnaam' zijn verplicht")
        if not wtype:
            raise HTTPException(400, f"Rij {i}: 'type' is verplicht")

        try:
            qty = int(aantal)
            if qty < 1:
                raise ValueError
        except ValueError:
            raise HTTPException(400, f"Rij {i}: ongeldig aantal '{aantal}'")

        vintage = None
        if jaargang:
            try:
                vintage = int(jaargang)
            except ValueError:
                raise HTTPException(400, f"Rij {i}: ongeldige jaargang '{jaargang}'")

        rows.append({
            "producer": producent,
            "wine_name": wijnnaam,
            "wine_type": wtype,
            "vintage": vintage,
            "volume": volume or "0.75L",
            "quantity": qty,
        })

    if not rows:
        raise HTTPException(400, "CSV bevat geen regels")
    return rows


def _parse_excel(content: bytes) -> list[dict]:
    """Parse Excel (.xlsx) content."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(400, "Excel bestand bevat geen werkblad")

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise HTTPException(400, "Excel bestand is leeg")

    # Normalize headers
    headers = [str(h).strip().lower() if h else "" for h in header]
    col_map = {h: i for i, h in enumerate(headers)}

    required = ["producent", "wijnnaam", "type"]
    for req in required:
        if req not in col_map:
            raise HTTPException(400, f"Kolom '{req}' niet gevonden. Verwacht: producent, wijnnaam, type, jaargang, volume, aantal")

    rows = []
    for i, row in enumerate(rows_iter, start=2):
        def cell(name: str) -> str:
            idx = col_map.get(name)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        producent = cell("producent")
        wijnnaam = cell("wijnnaam")
        wtype = cell("type")

        if not producent or not wijnnaam:
            continue  # skip empty rows

        if not wtype:
            raise HTTPException(400, f"Rij {i}: 'type' is verplicht")

        jaargang = cell("jaargang")
        volume = cell("volume") or "0.75L"
        aantal = cell("aantal") or "1"

        try:
            qty = int(float(aantal))
            if qty < 1:
                raise ValueError
        except ValueError:
            raise HTTPException(400, f"Rij {i}: ongeldig aantal '{aantal}'")

        vintage = None
        if jaargang:
            try:
                vintage = int(float(jaargang))
            except ValueError:
                raise HTTPException(400, f"Rij {i}: ongeldige jaargang '{jaargang}'")

        rows.append({
            "producer": producent,
            "wine_name": wijnnaam,
            "wine_type": wtype,
            "vintage": vintage,
            "volume": volume,
            "quantity": qty,
        })

    wb.close()

    if not rows:
        raise HTTPException(400, "Excel bestand bevat geen regels")
    return rows


@router.post("/import", response_model=OrderImportResult)
async def import_order(
    file: UploadFile,
    order_number: str = "",
    customer_name: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(require_product_manager),
):
    """Import an order from CSV or Excel file.

    Expected columns: producent, wijnnaam, type, jaargang, volume, aantal.
    Existing SKUs are matched by generated sku_code, new SKUs are created without images.
    """
    content = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".xlsx"):
        rows = _parse_excel(content)
    else:
        rows = _parse_csv(content)

    # Generate order number if not provided
    if not order_number:
        order_number = f"ORD-{uuid.uuid4().hex[:8].upper()}"
    if not customer_name:
        customer_name = "Onbekend"

    existing = db.query(Order).filter(Order.order_number == order_number).first()
    if existing:
        raise HTTPException(400, f"Ordernummer '{order_number}' bestaat al")

    order = Order(order_number=order_number, customer_name=customer_name)
    db.add(order)
    db.flush()

    new_skus = []
    existing_skus = []

    for row in rows:
        sku_code = generate_sku_code(
            row["producer"], row["wine_name"], row["wine_type"], row["vintage"], row["volume"]
        )
        sku = db.query(SKU).filter(SKU.sku_code == sku_code).first()

        if sku:
            existing_skus.append(sku)
        else:
            name = generate_display_name(row["producer"], row["wine_name"], row["vintage"], row["volume"])
            sku = SKU(
                sku_code=sku_code,
                name=name,
                producer=row["producer"],
                wine_name=row["wine_name"],
                wine_type=row["wine_type"],
                vintage=row["vintage"],
                volume=row["volume"],
            )
            db.add(sku)
            db.flush()
            new_skus.append(sku)

        # Check if a line for this SKU already exists on this order
        existing_line = (
            db.query(OrderLine)
            .filter(OrderLine.order_id == order.id, OrderLine.sku_id == sku.id)
            .first()
        )
        if existing_line:
            existing_line.quantity += row["quantity"]
        else:
            db.add(OrderLine(order_id=order.id, sku_id=sku.id, quantity=row["quantity"]))

    db.commit()
    db.refresh(order)

    publish_event(
        "order_imported",
        details={
            "order_number": order.order_number,
            "customer_name": order.customer_name,
            "total_lines": len(order.lines),
            "new_skus": [s.sku_code for s in new_skus],
        },
        user=user,
        resource_type="order",
        resource_id=order.id,
    )

    return OrderImportResult(
        order=_order_to_response(order),
        new_skus=[_sku_to_response(s) for s in new_skus],
        existing_skus=[_sku_to_response(s) for s in existing_skus],
    )


@router.get("", response_model=list[OrderResponse])
def list_orders(
    status: str | None = None,
    db: Session = Depends(get_db),
):
    query = db.query(Order)
    if status:
        query = query.filter(Order.status == status)
    orders = query.order_by(Order.created_at.desc()).all()
    return [_order_to_response(o) for o in orders]


@router.get("/{order_id}", response_model=OrderResponse)
def get_order(order_id: int, db: Session = Depends(get_db)):
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(404, "Order niet gevonden")
    return _order_to_response(order)


@router.delete("/{order_id}", status_code=204)
def delete_order(
    order_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_product_manager),
):
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(404, "Order niet gevonden")
    order_number = order.order_number
    db.delete(order)
    db.commit()
    publish_event(
        "order_deleted",
        details={"order_number": order_number},
        user=user,
        resource_type="order",
        resource_id=order_id,
    )


@router.post("/{order_id}/activate", response_model=OrderResponse)
def activate_order(
    order_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_product_manager),
):
    """Activate an order. All SKUs must have at least one reference image."""
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(404, "Order niet gevonden")
    if order.status != "draft":
        raise HTTPException(400, f"Order is al '{order.status}'")

    # Check every SKU has at least one reference image
    missing = []
    for line in order.lines:
        if len(line.sku.reference_images) == 0:
            missing.append(line.sku.sku_code)
    if missing:
        raise HTTPException(
            400,
            f"Kan niet activeren: {len(missing)} SKU('s) zonder referentiebeeld: {', '.join(missing[:5])}"
        )

    order.status = "active"
    db.commit()
    db.refresh(order)

    publish_event(
        "order_activated",
        details={"order_number": order.order_number},
        user=user,
        resource_type="order",
        resource_id=order.id,
    )

    return _order_to_response(order)


@router.post("/{order_id}/scan", response_model=ScanResult)
async def scan_box(
    order_id: int,
    file: UploadFile,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Scan a box for this order. Matches vision result against order lines."""
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(404, "Order niet gevonden")
    if order.status != "active":
        raise HTTPException(400, "Order is niet actief")

    image_bytes = await file.read()

    # Save scan image
    scan_dir = os.path.join(settings.upload_dir, "scans")
    os.makedirs(scan_dir, exist_ok=True)
    scan_path = os.path.join(scan_dir, f"{uuid.uuid4().hex}.jpg")
    with open(scan_path, "wb") as f:
        f.write(image_bytes)

    description, embedding = process_image(image_bytes)
    candidates = find_best_matches(db, embedding, top_n=5)

    matched_sku, confidence = None, 0.0
    if candidates and candidates[0][1] >= settings.match_threshold:
        matched_sku, confidence = candidates[0]

    # Check if matched SKU is in this order
    matched_line = None
    if matched_sku:
        for line in order.lines:
            if line.sku_id == matched_sku.id and line.scanned_quantity < line.quantity:
                matched_line = line
                break

    if matched_line:
        matched_line.scanned_quantity += 1
        db.commit()

        # Check if order is fully scanned
        all_done = all(l.scanned_quantity >= l.quantity for l in order.lines)
        if all_done:
            order.status = "completed"
            db.commit()

        total_boxes = sum(l.quantity for l in order.lines)
        scanned_boxes = sum(l.scanned_quantity for l in order.lines)

        publish_event(
            "order_scan",
            details={
                "order_number": order.order_number,
                "matched_sku_code": matched_sku.sku_code,
                "confidence": round(confidence, 4),
                "vision_description": description,
                "scanned_total": f"{scanned_boxes}/{total_boxes}",
                "completed": all_done,
            },
            user=user,
            resource_type="order",
            resource_id=order.id,
        )

        message = f"Zet op rolcontainer {order.customer_name}"
        if all_done:
            message = f"Order {order.order_number} compleet!"

        return ScanResult(
            matched=True,
            sku_code=matched_sku.sku_code,
            sku_name=matched_sku.name,
            confidence=confidence,
            order_line_id=matched_line.id,
            scanned_quantity=matched_line.scanned_quantity,
            total_quantity=matched_line.quantity,
            customer_name=order.customer_name,
            message=message,
        )

    # No match in this order
    publish_event(
        "order_scan",
        details={
            "order_number": order.order_number,
            "matched_sku_code": matched_sku.sku_code if matched_sku else None,
            "confidence": round(confidence, 4) if matched_sku else None,
            "vision_description": description,
            "in_order": False,
        },
        user=user,
        resource_type="order",
        resource_id=order.id,
    )

    if matched_sku:
        return ScanResult(
            matched=False,
            sku_code=matched_sku.sku_code,
            sku_name=matched_sku.name,
            confidence=confidence,
            customer_name=order.customer_name,
            message=f"Dit product ({matched_sku.name}) hoort niet bij deze order",
        )

    return ScanResult(
        matched=False,
        customer_name=order.customer_name,
        message="Product niet herkend",
    )
